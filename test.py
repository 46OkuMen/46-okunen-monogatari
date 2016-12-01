"""Hopelessly late 'unit tests' for the reinserter."""
# At this point I need to test various things about the data, not really the program itself.
# So this doesn't really count much for 'test coverage.' Oh well.

# For this case, unit tests are not useful - I need to check things about the data itself.
# Unit tests stop after one thing is wrong; I need a list of all things that are wrong.

import os
from utils import DUMP_XLS, POINTER_XLS, onscreen_length, DEST_PATH
from openpyxl import load_workbook
from rominfo import file_blocks, POINTER_CONSTANT, CREATURE_BLOCK, CREATURE_MAX_LENGTH, DIALOGUE_MAX_LENGTH, DAT_MAX_LENGTH, FULLSCREEN_MAX_LENGTH
from pointer_peek import word_at_offset, text_at_offset, text_with_pointer

def test_increasing_offsets():
    """Make sure the offsets are strictly increasing - so that no strings are mislabeled."""
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    sheets.remove('SINKA.DAT')
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):  # Skip the first row, it's just labels
            try:
                this_offset = int(row[0].value, 16)
            except TypeError:
                break
            if this_offset <= new_offset:
                print "In sheet %s, fix offset at row %s; %s < %s" % (sheet, index+2, hex(int(row[0].value, 16)), hex(new_offset))
            new_offset = this_offset

def test_substrings_of_earlier_strings():
    """
    The reinserter sometimes accidentally replaces something earlier in the file, causing hard-to-find pointer bugs.
    It does this when the japanese string that's been translated is a substring of an untranslated earlier string.
    If I know where this happens, I can fix it manually pretty easily.
    """
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    sheets.remove('SINKA.DAT')
    sheets.remove('SEND.DAT')
    for gamefile in file_blocks:
        ws = wb.get_sheet_by_name(gamefile)
        for (start, stop) in file_blocks[gamefile]:
            previous_untranslated_jp = []
            for index, row in enumerate(ws.rows[1:]):
                try:
                    this_offset = int(row[0].value, 16)
                except TypeError:
                    break
                if start <= this_offset <= stop:
                    jp_string = row[2].value 
                    if isinstance(jp_string, basestring):
                        for i in previous_untranslated_jp:    # Repeats (nametags, etc) are also substrings!
                            if jp_string in i[0] and row[4].value:     # If it's occurred before AND it's been translated:
                                print "%s, '%s' is a substring of the untranslated string at %s" % (gamefile, row[4].value, i[1])
                                # More informative way of doing this...? Terminal can't display jp text though.
                                break
                        # If it's not translated, add it to the list of things to check against later
                        if row[4].value is None:
                            previous_untranslated_jp.append((jp_string, row[0].value))

def test_duplicate_strings():
    """
    If multiple English strings are duplicates, we can reroute the pointers to just one and gain space.
    """

    # TODO: Remove nametags from consideration.
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    sheets.remove('SINKA.DAT')
    sheets.remove('SEND.DAT')
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        previous_en = []
        for index, row in enumerate(ws.rows[1:]):
            try:
                this_offset = int(row[0].value, 16)
            except TypeError:
                break
            en_string = row[4].value 
            if isinstance(en_string, basestring):
                for i in previous_en:
                    if en_string == i[0]:
                        print "%s, '%s' is a duplicate of a string at %s" % (sheet, row[4].value, i[1])
                        break
                previous_en.append((en_string, row[0].value))

def test_all_string_lengths():
    """Corasest string test. No string can be longer than 77."""
    wb = load_workbook(DUMP_XLS)
    sheets = wb.get_sheet_names()
    sheets.remove('ORIGINAL')
    sheets.remove('MISC TITLES')
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[4].value:
                print row[4].value
                if onscreen_length(row[4].value) > FULLSCREEN_MAX_LENGTH:
                    print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))


def test_creature_string_lengths():
    """No creature name can have a name longer than 21."""
    wb = load_workbook(DUMP_XLS)
    for sheet in ['ST1.EXE', 'ST2.EXE', 'ST3.EXE', 'ST4.EXE', 'ST5.EXE', 'ST6.EXE']:
        creature_lo, creature_hi = CREATURE_BLOCK[sheet]
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if int(row[0].value, 16) >= creature_lo and int(row[0].value, 16) <= creature_hi:
                if row[4].value:
                    assert onscreen_length(row[4].value) <= CREATURE_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)


def test_dat_string_lengths():
    """No encyclopedia string should be longer than 68."""
    wb = load_workbook(DUMP_XLS)
    for sheet in ['SEND.DAT', 'SINKA.DAT']:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[4].value:
                assert onscreen_length(row[4].value) <= DAT_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)


def test_game_string_lengths():
    """
    Wide onscreen strings can't be more than 68? characters.
    Ones not marked as wide can't be more than 42 characters.
    """
    wb = load_workbook(DUMP_XLS)
    for sheet in ['ST1.EXE', 'ST2.EXE', 'ST3.EXE', 'ST4.EXE', 'ST5.EXE', 'ST5S1.EXE',
                  'ST5S2.EXE', 'ST5S3.EXE', "ST6.EXE"]:
        ws = wb.get_sheet_by_name(sheet)
        new_offset = 0
        for index, row in enumerate(ws.rows[1:]):
            if row[7].value:
                if row[7].value == 'wide':
                    if onscreen_length(row[4].value) > FULLSCREEN_MAX_LENGTH:
                        try:
                            print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))
                        except UnicodeDecodeError:
                            print "%s %s: %s has length %s" % (sheet, row[0].value, 'some string', onscreen_length(row[4].value))
            else:
                if onscreen_length(row[4].value) > DIALOGUE_MAX_LENGTH:
                    try:
                        print "%s %s: %s has length %s" % (sheet, row[0].value, row[4].value, onscreen_length(row[4].value))
                    except UnicodeDecodeError:
                        print "%s %s: %s has length %s" % (sheet, row[0].value, 'some string', onscreen_length(row[4].value))

# make sure a blank translation sheet doesn't return overflow errors

def test_map_locations():
    """
    Pointers to .MAP files should return the same pointed text after reinsertion.
    If not, the game will probably crash when trying to load it.
    """
    wb = load_workbook(POINTER_XLS)
    ws = wb.get_sheet_by_name('Sheet1')
    for row in ws.rows[1:]:
        if row[3].value:
            if row[3].value.endswith('.MAP'):
                src_filename = row[0].value
                pointer_location = row[2].value
                expected_text = row[3].value

                assert expected_text == text_with_pointer(src_filename, pointer_location), "pointer %s has the wrong value: %s" % (pointer_location, text_with_pointer(src_filename, pointer_location))

def test_image_locations():
    """
    Pointers to .GDT files should return the same pointed text after reinsertion.
    If not, the game will probably crash when trying to load it.
    """
    wb = load_workbook(POINTER_XLS)
    ws = wb.get_sheet_by_name('Sheet1')
    for row in ws.rows[1:]:
        if row[3].value:
            if row[3].value.endswith('.GDT'):
                src_filename = row[0].value
                pointer_location = row[2].value
                expected_text = row[3].value

                assert expected_text == text_with_pointer(src_filename, pointer_location), "pointer %s has the wrong value: %s" % (pointer_location, text_with_pointer(src_filename, pointer_location))

def test_merged_pointers():
    """Assert that two originally seperate pointers now point to the same text location.
    """
    from rominfo import POINTER_ABSORB
    # So, one issue is that the "Cancel" strings are the first one in their block, which means
    # no diff is ever calculated for them.

    # I need to do this as a part of the reinserter, not just the pointer sheet.
    # When absorbing a pointer, set its value equal before adjusting both.
    for (filename, a), b_text_offset in POINTER_ABSORB.iteritems():
        gamefile_path = os.path.join(DEST_PATH, filename)
        first = word_at_offset(gamefile_path, a) + POINTER_CONSTANT[filename]
        assert first == b_text_offset, '%s: %s, %s' % (hex(a), hex(first + POINTER_CONSTANT[filename]), hex(b_text_offset))

def test_block_increasing_offsets():
    for f in file_blocks:
        offsets = list(sum(f, ()))
        print offsets

if __name__ == '__main__':
    test_increasing_offsets()
    #test_game_string_lengths()
    #test_substrings_of_earlier_strings()
    #est_duplicate_strings()
    test_map_locations()
    test_image_locations()
    #test_merged_pointers()
    test_block_increasing_offsets()