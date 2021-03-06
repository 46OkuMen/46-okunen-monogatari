"""Classes for the 46 Okunen Monogatari text reinserter."""

from __future__ import division
import os
from binascii import unhexlify
from math import floor

from openpyxl import load_workbook

from utils import pack, unpack, file_to_hex_string, DUMP_XLS, POINTER_XLS, SRC_PATH, DEST_PATH
from utils import sjis_to_hex_string, ascii_to_hex_string
from utils import onscreen_length
from rominfo import file_blocks, file_location, file_length, POINTER_CONSTANT, POINTER_ABSORB, POINTER_CHANGE
from rominfo import SPARE_BLOCK, OTHER_SPARE_BLOCK, CREATURE_BLOCK
from rominfo import DAT_MAX_LENGTH, FULLSCREEN_MAX_LENGTH, DIALOGUE_MAX_LENGTH, MEDIUM_MAX_LENGTH

from pointer_peek import text_at_offset, word_at_offset

class Disk(object):
    """The main .FDI file for a PC-98 game. Disks have the properties:

    Attributes:
        src_path: A string with the path to the original .FDI.
        dest_path: A string with the path to the patched .FDI.
        dest_dir: A string to the path of the containing folder of the patched .FDI.
        original_romstring: A hex string of the entire file. Don't change it!
        romstring: A hex string of the entire file. Change this!
        gamefiles: EXEFile and DATFile objects identified in files_to_translate.

    Methods:
        translate(): Perform reinsertion and translation of all files.
        write(): Write the patched bytes to dest_path.
        """ 

    def __init__(self, src_path, dest_path, files_to_translate):
        self.src_path = src_path
        self.dest_path = dest_path

        self.dest_dir = os.path.abspath(os.path.join(dest_path, os.pardir))

        self.original_romstring = file_to_hex_string(src_path)
        self.romstring = "" + self.original_romstring

        self.dump_excel = DumpExcel(DUMP_XLS)
        self.pointer_excel = PointerExcel(POINTER_XLS)

        # If I want to hardcode the total strings here, it's 4,815 (9/7/16)
        # wait - when I don't hardcode it, it's 5,240?? (Is that just the numbers in the .dat files?)
        # # TODO: Determine number of strings in whole game.
        #self.total_strings = 4815
        self.total_strings = 0
        self.translated_strings = 0

        self.gamefiles = []
        for filename in files_to_translate:
            if filename.endswith('.EXE'):
                exefile = EXEFile(self, filename)
                self.total_strings += exefile.total_strings
                self.translated_strings += exefile.translated_strings
                self.gamefiles.append(exefile)
            elif filename.endswith('.DAT'):
                datfile = DATFile(self, filename)
                self.total_strings += datfile.total_strings
                self.translated_strings += datfile.translated_strings
                self.gamefiles.append(datfile)

        
    def translate(self):
        """Perform translation and reinsertion."""
        for gamefile in self.gamefiles:
            if gamefile.filename.endswith('DAT'):
                gamefile.translate()
                gamefile.write()
                gamefile.report_progress()
            else:
                for block in gamefile.blocks:
                    block.edit_text()
                try:
                    gamefile.creature_block.edit_text()
                except AttributeError:
                    # Doesn't have a creature block. Don't worry about it.
                    pass
                gamefile.move_overflow()
                gamefile.incorporate()

                gamefile.write()
                gamefile.report_progress()
        try:
            self.write()
        except IOError:
            raw_input("You have the game open; close it and hit Enter to continue")
            self.write()
        self.report_progress()

    def typeset(self, files):
        """Perform typesetting."""
        for filename in files:
            gamefile = [g for g in self.gamefiles if g.filename == filename][0]
            for block in gamefile.blocks:
                block.typeset()
            gamefile.incorporate()
            #print "typeset", filename
            gamefile.write()
        self.write()

    def write(self):
        """Write the patched bytes to a new FDI."""
        data = unhexlify(self.romstring)
        with open(self.dest_path, 'wb') as fileopen:
            fileopen.write(data)

    def report_progress(self):
        """Calculate and print the progress made in translating this file."""

        percentage = int(floor((self.translated_strings / self.total_strings * 100)))
        print 'E.V.O.: The Theory of Evolution', str(percentage), "% complete",
        print "(%s / %s)\n" % (self.translated_strings, self.total_strings)


class Gamefile(object):
    """
    Any file on the disk targeted for reinsertion.

    Attributes:
        filename: String with filename and extension.
        disk: The Disk object containing the file.
        location: Int, starting offset of the file in DiskA.FDI. (Not valid for HDI.)
        length: Int, length in bytes of the file.
        original_filestring: Hex string of the untranslated string.
        filestring: Hex string of the file; gets edited during reinsertion.
        blocks: List of Block objects belonging to the file.
        total_strings: Number of Japanese strings in the file.
        translated_strings: Number of replacements made.

    Methods:
        incorporate(): Reinsert this translated file into its disk.
        write(): Write this file independently to the patched directory.
        report_progress(): Print the file's percent completion.
    """

    def __init__(self, disk, filename):
        self.filename = filename
        self.disk = disk
        self.location = file_location[filename]
        self.length = file_length[filename]

        self.original_filestring = file_to_hex_string(disk.src_path, self.location, self.length)
        self.filestring = "" + self.original_filestring

        self.blocks = []
        for block in file_blocks[self.filename]:
            self.blocks.append(Block(self, block))

        self.total_strings = 0
        self.translated_strings = 0
        self.spaces = 0
        for block in self.blocks:
            for trans in block.translations:
                if isinstance(trans.japanese, long):
                    # Skip the numbers in .DAT files, they're boring
                    continue
                self.total_strings += 1
                if trans.english:
                    self.translated_strings += 1

    def incorporate(self):
        """Add the edited file to the Disk in the original's place."""
        for b in self.blocks:
            b.incorporate()

        try:
            i = self.disk.romstring.index(self.original_filestring)
            self.disk.romstring = self.disk.romstring.replace(self.original_filestring, self.filestring, 1)
        except ValueError:
            print "Can't find that file, try replacing it instead"
            # Do pointers get updated in the filestring as well?
            self.disk.romstring = self.disk.romstring.replace(self.disk.romstring[self.location*2:(self.location+self.length)*2], self.filestring)

        # Set the "original filestring" to the current one to incorporate again after typesetting.
        self.original_filestring = self.filestring

    def write(self):
        """Write the new data to an independent file for later inspection."""
        data = unhexlify(self.filestring)
        dest_path = os.path.join(self.disk.dest_dir, self.filename)
        with open(dest_path, 'wb') as fileopen:
            fileopen.write(data)

    def get_spaces(self):
        """Count the number of characters freed from SJIS spaces in the file."""
        for b in self.blocks:
            for t in b.translations:
                self.spaces += t.spaces
        return self.spaces*2

    def block_at(self, offset):
        """Return the block object of a given offset in this file."""
        for block in self.blocks:
            lo, hi = block.start, block.stop
            if (offset >= lo) and (offset < hi):
                return block

    def report_progress(self):
        """Calculate and print the progress made in translating this file."""
        percentage = int(floor((self.translated_strings / self.total_strings * 100)))
        print self.filename, str(percentage), "% complete",
        print "(%s / %s)" % (self.translated_strings, self.total_strings),
        print "%s characters of space taken from SJIS spaces\n" % self.get_spaces()

    def __repr__(self):
        return self.filename


class EXEFile(Gamefile):
    """An executable gamefile. Needs to deal with pointers and blocks and overflow.

    Attributes:
        pointer_constant: File-specific constant added to pointer values to retrive text location.
        pointers: A dict of {text_location: Pointer} pairs.
        spare_block: A Block identified as expendable; overflow text can be placed here.
        creature_block: A Block with different padding rules.
        overflows: A list of Overflows which will get moved and repointed elsewhere.

    Methods:
        edit_pointers_in_range: Adjust the pointers which point between (lo, hi) with a given diff.
        get_pointers: Retrive a dict of Pointer objects.
        move_overflow: Move overflow to the spare block and adjust pointers.
        (change_starting_map: Cheat and change the beginning-of-chapter spawn point. (Not implemented yet))
         """

    def __init__(self, disk, filename):
        Gamefile.__init__(self, disk, filename)
        self.disk = disk
        self.pointer_constant = POINTER_CONSTANT[filename]

        # Look for a spare block and designate it as such.
        self.spares = []
        self.spare_block = None
        self.other_spare_block = None

        try:
            spare_start, spare_stop = SPARE_BLOCK[self.filename]
            self.spares.append((spare_start, spare_stop))
            for block in self.blocks:
                if block.start == spare_start:
                    self.spare_block = block
        except KeyError:
            self.spare_block = None

        try:
            other_start, other_stop = OTHER_SPARE_BLOCK[self.filename]
            self.spares.append((other_start, other_stop))
            for block in self.blocks:
                if block.start == other_start:
                    self.other_spare_block = block
        except KeyError:
            self.other_spare_block = None

        # Then look for a creature block and designate it as such.
        try:
            creature_start, creature_stop = CREATURE_BLOCK[self.filename]
            for block in self.blocks:
                if block.start == creature_start:
                    self.creature_block = CreatureBlock(self, (creature_start, creature_stop))
                    self.blocks.remove(block)
        except KeyError:
            self.creature_block = None

        self.pointers = self._get_pointers()

        self.overflows = []

        self.total_strings = 0
        self.translated_strings = 0

        for block in self.blocks:
            for trans in block.translations:
                self.total_strings += 1
                if trans.english:
                    self.translated_strings += 1

        if self.creature_block:
            for c in self.creature_block.translations:
                self.total_strings += 1
                if c.english:
                    self.translated_strings += 1

    def edit_pointers_in_range(self, (start, stop), diff):
        """Edit all the pointers between two file offsets."""
        if diff != 0:
            for offset in [p for p in range(start+1, stop+1) if p in self.pointers]:
                for ptr in self.pointers[offset]:
                    ptr.edit(diff)

    def _get_pointers(self):
        """Retrieve all relevant pointers from the pointer sheet."""
        # The pointers in a given file will never change. But the pointer objects and values will...
        result = self.disk.pointer_excel.get_pointers(self)
        return result

    def refresh_pointers(self):
        self.pointers = self.disk.pointer_excel.get_pointers(self, refresh=True)
        return self.pointers

    def move_overflow(self):
        """
        Move the overflows to the spares, then reroute the pointers.
        """
        if self.spare_block:
            self.spare_block.blockstring = ""
        if self.other_spare_block:
            self.other_spare_block.blockstring = ""

        # Want to try to put the largest overflows into the smallest containing spare, for best space usage.
        self.overflows.sort(key=lambda x: x.new_length)
        self.spares.sort(key=lambda x: x[1] - x[0])
        
        while len(self.overflows) > 0:
            #print "Overflow:", [p.new_length for p in self.overflows]
            #print "Spares:", [s[1] - s[0] for s in self.spares]
            overflow = self.overflows.pop()
            overflow_stored = False
            overflow_length = overflow.new_length

            #print "Trying to store %s with length %s" % (overflow, overflow_length)
            for i, s in enumerate(self.spares):
                if s[1] - s[0] >= overflow_length:
                    #print overflow, "should fit in", hex(s[0]), hex(s[1]), "since %s < %s" % (overflow.new_length, s[1]-s[0])
                    overflow.move(s[0])
                    self.spares[i] = s[0] + overflow_length, s[1]
                    self.spares.sort(key=lambda x: x[1] - x[0])
                    overflow_stored = True
                    break

            if not overflow_stored:
                raise Exception("That overflow didn't fit anywhere")

        if self.spare_block:
            excess = len(self.spare_block.blockstring)//2 - (self.spare_block.stop - self.spare_block.start)
            assert excess <= 0, "Spare block is %s too long" % (excess)

class DATFile(Gamefile):
    """A data gamefile. Doesn't have pointers or a fixed length, so it's much simpler."""

    def __init__(self, disk, filename):
        Gamefile.__init__(self, disk, filename)
        self.blocks = []
        # DATFiles just get one big DATBlock.
        for block in file_blocks[self.filename]:
            self.blocks.append(Block(self, block))

        self.src_path = os.path.join(self.disk.src_path, filename)

    def translate(self):
        """Replace all japanese strings with english ones.
        If only it were always that easy!"""
        # (Actually a little more complicated. Gotta replace [PAGE] with appropriate [SENLN]s...)
        lines_since_page_break = 0
        for trans in self.blocks[0].get_translations():
            if trans.english == "":
                continue

            if self.filename == 'SEND.DAT':
                trans.english = trans.send_typeset()
                if '[PAGE]' in trans.english:
                    lines_since_page_break += trans.english.count('\n') + 1
                    trans.english = trans.english.replace('[PAGE]', '[SENLN]'*(4-lines_since_page_break))
                    lines_since_page_break = 0
                elif '/*' in trans.english or '.GDT' in trans.english or trans.english.startswith('-'):
                    pass
                else:
                    lines_since_page_break += trans.english.count('\n') + 1

            en_bytestring = ascii_to_hex_string(trans.english)

            jp_bytestring = trans.jp_bytestring

            try:
                i = self.filestring.index(jp_bytestring)
            except ValueError:
                if jp_bytestring.endswith('0d0a'):
                    jp_bytestring = jp_bytestring[:-4]
                try:
                    i = self.filestring.index(jp_bytestring)
                except ValueError:
                    if jp_bytestring.endswith('0d0a'):
                        jp_bytestring = jp_bytestring[:-4]
                    i = self.filestring.index(jp_bytestring)
            self.filestring = self.filestring.replace(jp_bytestring, en_bytestring, 1)


class Block(object):
    """A text block.

    Attributes:
        gamefile: The EXEFile or DATFile object it belongs to.
        start = Beginning offset of the block.
        stop  = Ending offset of the block.
        original_blockstring: Hex string of entire block.
        blockstring: Hex string of entire block for editing.
        translations: List of Translation objects.
        """

    def __init__(self, gamefile, (start, stop)):
        self.gamefile = gamefile
        self.start = start
        self.stop = stop

        start_in_disk = start + self.gamefile.location
        self.original_blockstring = file_to_hex_string(self.gamefile.disk.src_path,
                                                       start_in_disk, (self.stop-self.start))
        self.blockstring = "" + self.original_blockstring
        self.translations = self.get_translations()

    def get_translations(self):
        """Grab all translations in this block."""
        excel = self.gamefile.disk.dump_excel
        return excel.get_translations(self)

    def get_pointers(self):
        file_pointers = self.gamefile.pointers
        block_pointers = [p for p in file_pointers if p >= self.start and p <= self.stop]
        block_pointers.sort()
        return block_pointers

    def overflow_location(self):
        """
        Find the first pointer that contains text that will overflow.
        """
        result = None
        diff = 0
        block_length = self.stop - self.start

        block_pointers = self.get_pointers()

        block_pointers.append(self.stop) # Should solve the problem of the last pointer not being considered (due to ranges)

        for i, ptr in enumerate(block_pointers):
            if i >= len(block_pointers)-1:
                break
            ptr_range = (ptr, block_pointers[i+1])
            
            # Look for all translations located in the pointer range.
            translations = [t for t in self.translations if t.location >= ptr_range[0] and t.location < ptr_range[1]]

            for trans in translations:
                location_in_blockstring = (trans.location * 2) - self.start
                jp_bytestring, en_bytestring = trans.jp_bytestring, trans.en_bytestring

                if en_bytestring:
                    text_length = len(en_bytestring)//2
                    this_string_diff = (len(en_bytestring) - len(jp_bytestring)) // 2
                else:
                    text_length = len(jp_bytestring)//2
                    this_string_diff = 0

                new_text_offset = trans.location + text_length + diff # TODO: + this_string_diff?

                if new_text_offset >= self.stop:
                    while ptr + text_length + diff > self.stop:
                        # If the block is really tiny and overflows IMMEDIATELY, need to return asap.
                        if i == 0:
                            break
                        ptr = block_pointers[i-1]
                        i -= 1

                    return ptr

                diff += this_string_diff
                

    def edit_text(self):
        """Replace each japanese string in the block with the translated english string."""
        pointer_diff = 0
        previous_text_offset = self.start
        is_overflowing = False

        overflow_location = self.overflow_location()

        for trans in self.translations:
            jp_bytestring, en_bytestring = trans.jp_bytestring, trans.en_bytestring

            # jp_bytestring might include ASCII; if it's not found, try the ascii preserving method.
            try:
                #print jp_bytestring
                j = self.blockstring.index(jp_bytestring)
            except ValueError: # substring not found
                #print "using alt jp string"
                jp_bytestring = trans.jp_bytestring_alt

            if (trans.location >= overflow_location) and overflow_location:
                #print "overflow starting with", trans
                is_overflowing = True

                recent_pointer = overflow_location

                start_in_block = (recent_pointer - self.start)*2
                overflow_bytestring = self.original_blockstring[start_in_block:]
                # Store the start and end of the overflow bytestring, 
                # to make sure all pointers are adjusted in the range.
                overflow_lo, overflow_hi = recent_pointer, self.stop

                overflow_pointers = [p for p in self.get_pointers() if overflow_lo <= p <= overflow_hi]
                if self.stop not in overflow_pointers:
                    overflow_pointers.append(self.stop)

                for i, p in enumerate(overflow_pointers):
                    if i == len(overflow_pointers)-1:
                        break
                    next_p = overflow_pointers[i+1]
                    start_in_block = (p - self.start)*2
                    stop_in_block = (next_p - self.start)*2

                    this_bytestring = self.original_blockstring[start_in_block:stop_in_block]
                    #print this_bytestring
                    #print "all pointers in overflow:", [hex(p) for p in overflow_pointers]
                    #print hex(p), hex(next_p)
                    this_overflow = Overflow(self.gamefile, (p, next_p), this_bytestring)
                    #print this_overflow 
                    self.gamefile.overflows.append(this_overflow)

            self.gamefile.edit_pointers_in_range((previous_text_offset, trans.location), pointer_diff)

            previous_text_offset = trans.location

            # "eng" is an ugly way of doing this. But we want to be able to do the thing where, if
            # it's overflowing, do a "translation" of the overflow bytestring into "".
            # If we set trans.english to "", it wouldn't get translated later.
            # So use a separate temporary "eng" variable which can be english or blank, but don't lose data.
            eng = trans.english
            if eng == "":
                # No replacement necessary - pointers are edited, so we're done here.
                continue

            if is_overflowing:
                # Then we want to blank the entire overflow bytestring.
                # So use the rest of the function already there to do that.
                eng = ""
                jp_bytestring = overflow_bytestring
                #print "jp bytestring is an overflow bytestring"

            # Recalculate in case it got altered due to overflow.
            en_bytestring = ascii_to_hex_string(eng)
            this_string_diff = (len(en_bytestring) - len(jp_bytestring)) // 2

            # Predict where the string should be, then start looking there.
            # Not really sure why these estimations are wrong sometimes...
            location_in_blockstring = ((pointer_diff + trans.location - self.start) * 2)
            old_slice = self.blockstring[location_in_blockstring:]

            try:
                i = old_slice.index(jp_bytestring)//2
            except ValueError:
                print hex(trans.location), trans.english
                #print "overflowing?", is_overflowing
                #print "looking for", jp_bytestring
                #print "in the string:", old_slice
                old_slice = self.blockstring
                i = old_slice.index(jp_bytestring)//2

            new_slice = old_slice.replace(jp_bytestring, en_bytestring, 1)

            self.blockstring = self.blockstring.replace(old_slice, new_slice, 1)

            pointer_diff += this_string_diff

            if is_overflowing:
                break

        self.identify_spares()

    def incorporate(self):
        """Write the new block to the source gamefile."""
        self.pad()

        start_in_file = self.start*2
        stop_in_file = self.stop*2

        self.gamefile.filestring = self.gamefile.filestring[:start_in_file] + self.blockstring + self.gamefile.filestring[stop_in_file:]
        assert len(self.gamefile.original_filestring) == len(self.gamefile.filestring)

    def identify_spares(self):
        """
        Determine how much space is left at the end of the block, and designate it for
        overflow collection.
        """
        block_diff = (len(self.blockstring) - len(self.original_blockstring)) // 2
        assert block_diff <= 0, 'The block %s is too long by %s' % (self, block_diff)
        if block_diff < 0:
            number_of_spaces = (-1)*block_diff
            padding_start = self.start + len(self.blockstring)//2
            padding_stop = self.start + len(self.blockstring)//2 + number_of_spaces

            self.gamefile.spares.append((padding_start, padding_stop))

    def pad(self):
        """Fill the remainder of the block with spaces. After identifying the spares."""
        block_diff = len(self.blockstring) - len(self.original_blockstring)
        assert block_diff <= 0, 'The block %s is too long by %s' % (self, block_diff)
        if block_diff < 0:
            number_of_spaces = ((-1)*block_diff)//2
            self.blockstring += '20' * number_of_spaces  # (ASCII space)

        assert len(self.original_blockstring) == len(self.blockstring)

    def typeset(self):
        """Typeset the block's text by pointer."""

        pointer_diff = 0
        for p in self.get_pointers():
            this_pointer = self.gamefile.pointers[p][0]

            # Hmm. Make sure to update pointers' text location even when they overflow/change blocks and such...
            # I may as well do something to just reload all the pointer locations, maybe using word_at_offset.
            # It's a huge pain to mess with the pointer functions in a way that doesn't break reinsertion...
            #print self, hex(this_pointer.text_location)
            #block_of_pointer = self.gamefile.block_at(this_pointer.text_location)
            #assert block_of_pointer == self, "%s %s" % (block_of_pointer, self)

            # Don't try to typeset stuff that has no real text in it.
            # That would break things like NPC movement code...
            if len(this_pointer.translations) > 0:
                this_pointer.typeset()
                pointer_diff += 0

    def __repr__(self):
        return "(%s, %s)" % (hex(self.start), hex(self.stop))


class CreatureBlock(Block):
    """
    The creature block has all the names of creatures in that file, plus some stat info.
    There's a consistent length to each creature's entry in this block which can't be altered.
    If you pad the creature's name with spaces, the spaces show up in gameplay text...
    But if you pad it with 00 bytes, everything is fine!
    So no pointers need to be adjusted, you just need to pad the strings.
    So the edit_text() method is a lot simpler for a CreatureBlock.
    """

    def edit_text(self):
        for trans in self.translations:
            jp_bytestring = trans.jp_bytestring
            en_bytestring = trans.en_bytestring

            this_string_diff = (len(en_bytestring) - len(jp_bytestring)) // 2

            if this_string_diff <= 0:
                en_bytestring += "00"*(this_string_diff*(-1))
            else:
                jp_bytestring += "00"*(this_string_diff)

            this_string_diff = (len(en_bytestring) - len(jp_bytestring)) // 2
            assert this_string_diff == 0, 'creature diff not 0'

            location_in_blockstring = (trans.location - self.start) * 2
            old_slice = self.blockstring[location_in_blockstring:]

            try:
                i = old_slice.index(jp_bytestring)//2
            except ValueError:
                old_slice = self.blockstring
                i = old_slice.index(jp_bytestring)//2

            new_slice = old_slice.replace(jp_bytestring, en_bytestring, 1)

            self.blockstring = self.blockstring.replace(old_slice, new_slice, 1)

        self.incorporate()


class Translation(object):
    """Has an offset, a SJIS japanese string, and an ASCII english string."""
    def __init__(self, block, location, japanese, english, width, page_break_suffix=''):
        self.location = location
        self.block = block
        self.japanese = japanese
        self.english = english
        self.max_width = width

        self.location_in_blockstring = (location - block.start) * 2

        self.jp_bytestring = sjis_to_hex_string(japanese)
        self.en_bytestring = ascii_to_hex_string(english)

        self.jp_bytestring_alt = sjis_to_hex_string(japanese, preserve_spaces=True)

        self.spaces = 0
        if isinstance(block.gamefile, EXEFile) or block.gamefile.filename == 'SEND.DAT':
            self.integrate_spaces()
            self.add_padding()

        self.page_break_suffix = page_break_suffix

    def integrate_spaces(self):
        """
        All second and third lines of dialogue are prepended with an SJIS space (0x8140).
        To save space, we get rid of these.
        """
        # first, remove the SJIS spaces that are already prepended.
        while self.jp_bytestring[0:4] == '8140':
            self.jp_bytestring = self.jp_bytestring[4:]

        scan = 0
        loc = self.location_in_blockstring
        if self.block.gamefile.filename == 'SEND.DAT':
            scan = 4
            snippet_right_before = self.block.blockstring[loc-scan:loc-scan+4]
            while snippet_right_before == '8140':
                self.jp_bytestring = '8140' + self.jp_bytestring
                self.jp_bytestring_alt = '8140' + self.jp_bytestring_alt
                self.spaces += 1

                scan += 4
                snippet_right_before = self.block.blockstring[loc-scan:loc-scan+4]
        else:
            snippet_right_before = self.block.blockstring[loc:loc+4]
            while snippet_right_before == '8140':
                self.jp_bytestring = '8140' + self.jp_bytestring
                self.jp_bytestring_alt = '8140' + self.jp_bytestring_alt
                self.spaces += 1

                scan += 4
                snippet_right_before = self.block.blockstring[loc+scan:loc+scan+4]

        # If there's one SJIS space in the JP version, it might be indented dialogue.
        if scan == 4:
            if self.block.gamefile.filename not in ('SEND.DAT', 'OPENING.EXE', 'ENDING.EXE') and "EVO" not in self.english and "Text Speed" not in self.english and not self.english.startswith('"') and not self.english.startswith('[SPLIT]'):
                #try:
                #    print self, "has one indentation"
                #except UnicodeEncodeError:
                #    print "some string", "has one indentation"z
                self.english = " " + self.english
                self.en_bytestring = '20' + self.en_bytestring

        #if scan > 4:
        #    print "%s %s: %s spaces" % (self.block.gamefile, hex(self.location), scan//4)

    def add_padding(self):
        """Add a single space if the string will need typesetting."""
        if onscreen_length(self.english) > 2*self.max_width:
            self.english += ' ' + ' '
            self.en_bytestring += '2020'
        elif onscreen_length(self.english) > self.max_width:
            self.english += ' '
            self.en_bytestring += '20'
        #elif onscreen_length(self.english) == self.max_width:
        #    self.japanese += "\n"
        #    self.jp_bytestring += '0a'

    def send_typeset(self):
        """
        Typeset a simple DAT string.
        No pointer-editing or length checking beyond 2 lines.
        Only aware of the current translation; can't prepend excess to the next translation.
        """
        if isinstance(self.english, long):
            return self.english

        if onscreen_length(self.english) > DAT_MAX_LENGTH:
            words = self.english.split(' ')
            lines = []
            line = ''

            while words:
                while onscreen_length(line) <= DAT_MAX_LENGTH and words:
                    if onscreen_length(line + " " + words[0]) <= DAT_MAX_LENGTH:
                        if len(line) > 0:
                            line += " "
                        line += words.pop(0)
                    else:
                        line = line.rstrip(' ')
                        lines.append(line)
                        line = ''
                        break
            if line:
                lines.append(line)
            if len(lines) > 1:
                print self
            return '\r\n'.join(lines)
        else:
            return self.english

    def __repr__(self):
        return hex(self.location) + " " + self.english


class Pointer(object):
    """
    A pointer. Found in EXEFiles outside of Blocks. They can be edited with edit(diff).
    """
    def __init__(self, gamefile, pointer_location, text_location):
        self.gamefile = gamefile
        self.location = pointer_location
        self.text_location = text_location

        self.old_value = text_location - gamefile.pointer_constant
        old_bytes = pack(self.old_value)
        self.old_bytestring = "{:02x}".format(old_bytes[0]) + "{:02x}".format(old_bytes[1])
       
        # the location of the translation relative to the pointer comes into play here...
        # the original text location, minus the extra spaces, plus the length of its text
        self.text_location_stop = self._true_location() + len(self.jp_text())

        self.translations = self.get_translations()

        self.original_translations = self.get_translations(original_location=True)

        if self.original_translations:
            self.max_width = self.original_translations[0].max_width
        else:
            self.max_width = 1000

        self.new_text_location = self.text_location

    def get_translations(self, original_location=False):
        result = []
        for b in self.gamefile.blocks:
            if original_location:
                result += [t for t in b.translations if self._true_location() <= t.location < self.text_location_stop]
            else:
                result += [t for t in b.translations if self.text_location <= t.location < self.text_location_stop]
        return result

    def edit(self, diff):
        """Adjusts the pointer by diff, and writes the new value to the gamefile."""
        #print "editing %s with diff %s" % (self, diff)
        if diff != 0:
            location_in_string = self.location*2
            new_value = self.old_value + diff
            new_bytes = pack(new_value)
            new_bytestring = "{:02x}".format(new_bytes[0]) + "{:02x}".format(new_bytes[1])

            string_before = self.gamefile.filestring[0:location_in_string]
            string_after = self.gamefile.filestring[location_in_string+4:]
                
            self.gamefile.filestring = string_before + new_bytestring + string_after

            self.new_text_location += diff

    def _true_location(self):
        """
        I'll let you in on a little secret: "text_location" isn't where the
        pointer points, usually. That's just where the translation was picked up.
        """
        # DANGER: I changed this to look at the SRC path instead of DEST.
        gamefile_path = os.path.join(SRC_PATH, self.gamefile.filename)
        pointer_value = word_at_offset(gamefile_path, self.location)
        return pointer_value + POINTER_CONSTANT[self.gamefile.filename]

    def jp_text(self):
        """
        Get what the pointer points to, ending at the END byte (00).
        """
        gamefile_path = os.path.join(SRC_PATH, self.gamefile.filename)
        pointer_value = word_at_offset(gamefile_path, self.location)
        pointer_location = pointer_value + POINTER_CONSTANT[self.gamefile.filename]

        #print gamefile_path, hex(pointer_location)
        result = text_at_offset(gamefile_path, pointer_location)

        # Sometimes there are pointers to control code right before an END.
        # Look a bit further in these cases.
        if len(result) < 2:
            #print "super short text; let's go a little further"
            result = text_at_offset(gamefile_path, pointer_location+2)
        return result


    def text(self, go_until_wait=False):
        """
        Get what the pointer points to, ending at the END byte (00).
        """
        gamefile_path = os.path.join(DEST_PATH, self.gamefile.filename)
        pointer_value = word_at_offset(gamefile_path, self.location)
        pointer_location = pointer_value + POINTER_CONSTANT[self.gamefile.filename]

        if go_until_wait:
            result = text_at_offset(self.gamefile, pointer_location, go_until_wait=True)
        else:
            result = text_at_offset(self.gamefile, pointer_location)

        # Sometimes there are pointers to control code right before an END.
        # Look a bit further in these cases.
        if len(result) < 2:
            if go_until_wait:
                result = text_at_offset(self.gamefile, pointer_location+2, go_until_wait=True)
            else:
                result = text_at_offset(self.gamefile, pointer_location+2)
        return result

    def typeset(self):
        """
        Find all the newlines in the pointer, then move them around.
        """
        original_text = self.text()

        if original_text.isspace() or "Cancel" in original_text:
            return None

        # Bad stuff happens if this method gets its hands on menu options
        if "EVO" in original_text or 'Text Speed' in original_text or 'Is this alright?' in original_text:
            return None

        textlines = original_text.splitlines()

        if len(textlines) > 8:
            # Probably a pointer table
            return None

        # Ignore stuff that's manually indented a lot
        if original_text.startswith('    '):
            return None

        spill = False
        for t in textlines:
            if onscreen_length(t) >= self.max_width:
                spill = True
        if not spill:
            #print "unchanged"
            #print original_text
            return None

        try:
            initial_newline = original_text[0] == '\n'
            final_newline = original_text[-1] == '\n'
            #final_double_newline = original_text[-2] == '\n'
        except IndexError:
            initial_newline = False
            final_newline = False
            #final_double_newline = False

        is_dialogue = False
        for t in textlines:
            naked_start = t.strip('\x16').strip(" ").strip("!")
            naked_stop = t.strip(" ").strip("\n").strip("\x13")
            if naked_start.startswith("'") or naked_start.startswith('"') or naked_stop.endswith("'") or naked_stop.endswith('"'):
                is_dialogue = True

        try:
            final_ln_wait = original_text[-2:] ==      '\n\x13'
            final_ln_wait_ln = original_text[-3:] == '\n\x13\n'
            final_wait_ln_ln = original_text[-3:] == '\x13\n\n'
            if final_ln_wait:
                textlines.pop(-1)
                textlines[-1] = textlines[-1].rstrip('\x13').strip(' ')
            if final_ln_wait_ln:
                textlines.pop(-1)
            if final_wait_ln_ln:
                textlines.pop(-1)
                textlines[-1] = textlines[-1].rstrip('\x13')
        except IndexError:
            final_ln_wait = False
            final_ln_wait_ln = False
            final_wait_ln_ln = False

        for i, line in enumerate(textlines):
            if onscreen_length(line) > self.max_width:
                if i == len(textlines) - 1:
                    joinedlines = line
                else:
                    joinedlines = line + " " + textlines[i+1]
                words = joinedlines.split(' ')
                firstline = ''
                while onscreen_length(firstline + " ") <= self.max_width:
                    if onscreen_length(firstline) + onscreen_length(words[0]) < self.max_width:
                        # Only add a space if it's not empty to begin with.
                        if len(firstline) > 0:
                            firstline += " "
                        firstline += words.pop(0)
                    else:
                        firstline = firstline.rstrip(" ")
                        break
                secondline = ' '.join(words)

                # If it's dialogue, doesn't already have a space, and doesn't have asterisks around it, indent.
                if is_dialogue and not secondline.startswith(" ") and not (secondline.startswith("*") and secondline.endswith("*")) and not secondline.isspace():
                    secondline = ' ' + secondline

                # The double space means it was indented by the reinserter, but the line beginning changed.
                if is_dialogue and '   ' in secondline:
                    secondline = secondline.replace('   ', ' ', 1)
                    secondline = ' ' + secondline

                if is_dialogue and '  ' in secondline:
                    secondline = secondline.replace('  ', ' ', 1)

                if i == len(textlines) - 1:
                    textlines.append('')
                
                textlines[i], textlines[i+1] = firstline, secondline

        if textlines[-1].endswith(" "):
            textlines[-1] = textlines[-1].rstrip()

        if is_dialogue:
            for i, t in enumerate(textlines):
                #if t.startswith('"') or t.startswith("*")
                if not (t.startswith('"') or t.startswith(' ') or t.isspace() or t == "" or i == 0 or t.startswith("*")):
                    textlines[i] = " " + textlines[i]

                if t.strip(" ").startswith("*"):
                    textlines[i] = textlines[i].lstrip(" ")

                if t.strip('\x13').strip('\x16').startswith(' "'):
                    textlines[i] = textlines[i].replace(' "', '"', 1)

        if textlines[-1].isspace() or len(textlines[-1]) == 0:
            textlines.pop()

        new_text = '\n'.join(textlines)

        if final_ln_wait:
            new_text += '\n\x13'
        elif final_ln_wait_ln:
            new_text += "\n\x13\n"
        elif final_wait_ln_ln:
            new_text += "\x13\n\n"
        elif final_newline:
            new_text += "\n"

        # On the off chance that something is exactly the same length, remove the newline.
        textlines = new_text.split('\n')
        for t in textlines:
            if onscreen_length(t) == self.max_width:
                try:
                    i = new_text.index(t + "\n")
                except ValueError:
                    print "probably no newline after this:"
                    print t
                new_text = new_text.replace(t + "\n", t, 1)

        windows = new_text.split('\x13')
        for i, w in enumerate(windows):
            # ignore the first line if it's blank; it just causes the next line to be printed at the middle.
            # (Don't remove it from the actual window, just remove it from this copy of the window)
            w = w.lstrip('\n')

            # If there's more than 3 lines of text in the window, text will get scrolled offscreen.
            if w.count('\n') > 2 and w.endswith('\n'):
                windows[i] = w[:-1] + " "

        new_text = "\x13".join(windows)

        if original_text[-1] == '\n' and original_text[-2] != '\x13':
            newline_diff = new_text.count('\n') - original_text.count('\n')

        old_bytestring = ascii_to_hex_string(original_text)
        new_bytestring = ascii_to_hex_string(new_text)

        if old_bytestring != new_bytestring:

            if len(old_bytestring) != len(new_bytestring):
                diff = len(old_bytestring) - len(new_bytestring)
                # Salvaging tactics. Where can you take away/add spaces that won't break stuff?
                if '0a' in new_bytestring and diff == 2:
                    i = new_bytestring.index('0a')
                    new_bytestring = new_bytestring.replace('0a', '200a', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '13' in new_bytestring and diff == 2:
                    i = new_bytestring.index('13')
                    new_bytestring = new_bytestring.replace('13', '2013', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '0a0a' in new_bytestring and diff == 4:
                    i = new_bytestring[-10:].index('0a')
                    new_bytestring = new_bytestring.replace('0a0a', '0a20200a', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '2220' in new_bytestring[-10:] and diff == -2:
                    i = new_bytestring[-10:].index('2220')
                    # Gotta be careful of the end quote. Don't replace "<LN> with *.
                    new_bytestring = new_bytestring[:-10] + new_bytestring[-10:].replace('2220', '22', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '200a' in new_bytestring and diff == -2:
                    i = new_bytestring.index('200a')
                    new_bytestring = new_bytestring.replace('200a', '0a', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif new_bytestring[-10:].count('20') >= 2 and diff == -4:
                    new_bytestring = new_bytestring[:-10] + new_bytestring[-10:].replace('20', '', 2)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '20130a0a' in new_bytestring and diff == -2:
                    new_bytestring = new_bytestring.replace('20130a0a', '130a0a', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '2013' in new_bytestring and diff == 4:
                    i = new_bytestring.index('2013')
                    new_bytestring = new_bytestring.replace('2013', '20202013', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '0a130a' in new_bytestring and diff == 4:
                    i = new_bytestring.index('0a130a')
                    new_bytestring = new_bytestring.replace('0a130a', '0a2020130a', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                elif '2013' in new_bytestring and diff == -2:
                    i = new_bytestring.index('2013')
                    new_bytestring = new_bytestring.replace('2013', '13', 1)
                    diff = len(old_bytestring) - len(new_bytestring)
                    assert diff == 0
                else:
                    print "too hard to reinsert right now:"
                    print "diff", diff
                    print "is_dialogue", is_dialogue
                    print self.max_width
                    print new_text
                    print old_bytestring
                    print new_bytestring
                    return None

            diff = len(old_bytestring) - len(new_bytestring)
            #print self.max_width, is_dialogue
            #print new_text
            assert diff == 0
            try:
                i = self.gamefile.filestring.index(old_bytestring)
            except ValueError:
                print "Couldn't find it in the whole block for some reason:"
                print original_text
                return None

            b = self.gamefile.block_at(i//2)

            try:
                bi = b.blockstring.index(old_bytestring)
                b.blockstring = b.blockstring.replace(old_bytestring, new_bytestring, 1)
            except ValueError:
                print "Couldn't find it in that block for some reason"
            b.incorporate()

            return len(new_bytestring) - len(old_bytestring) # for pointer adjustment

    def __repr__(self):
        return "%s pointing to %s" % (hex(self.location), hex(self.text_location))


class DumpExcel(object):
    """
    Takes a dump excel path, and lets you get a block's translations from it.
    """
    def __init__(self, path):
        self.path = path
        self.workbook = load_workbook(self.path)

    def get_translations(self, block):
        """Get the translations for an EXE or DAT file."""
        # So they can make use of Translation() objects as well.
        trans = []    # translations[offset] = Translation()
        worksheet = self.workbook.get_sheet_by_name(block.gamefile.filename)

        lines_since_page_break = 0

        for row in worksheet.rows[1:]:  # Skip the first row, it's just labels
            try:
                offset = int(row[0].value, 16)
            except TypeError:
                # Either a blank line or a total value. Ignore it.
                break
            if block.start <= offset < block.stop:
                japanese = row[2].value
                english = row[4].value

                width = DIALOGUE_MAX_LENGTH
                if row[6].value:
                    if row[6].value == 'wide':
                        width = FULLSCREEN_MAX_LENGTH
                    elif row[6].value == 'medium':
                        width = MEDIUM_MAX_LENGTH

                if isinstance(japanese, long):
                    # Causes some encoding problems? Trying to skip them for now
                    continue

                # Yeah this is important - blank strings are None (non-iterable), so use "" instead.
                if not english:
                    english = ""

                # Convert [PAGE] control codes to the appropriate number of [SENLN] control codes.
                # Only necessary for SEND strings.
                if block.gamefile.filename == 'SEND.DAT':
                    if '[PAGE]' in japanese:
                        lines_since_page_break += 1
                        japanese = japanese.replace('[PAGE]', '')
                        # This will get replaced by the appropriate number of [SENLN]s by the typesetter.
                        #english += '[PAGE]'
                        lines_since_page_break = 0
                    elif '.GDT' in japanese or '/*' in japanese:
                        pass
                    else:
                        lines_since_page_break += 1
                    #assert lines_since_page_break <= 3
                trans.append(Translation(block, offset, japanese, english, width))
        return trans


class PointerExcel(object):
    """
    Takes a pointer dump excel path, and lets you grab the relevant pointers from it.
    """
    def __init__(self, path):
        self.path = path
        self.pointer_wb = load_workbook(self.path)
        self.pointer_sheet = self.pointer_wb.worksheets[0]

    def get_pointers(self, gamefile, refresh=False):
        """Retrieve all relevant pointers from the pointer sheet."""
        ptrs = {}

        try:
            spare_start, spare_stop = SPARE_BLOCK[gamefile.filename]
        except KeyError:
            spare_start, spare_stop = None, None

        for row in [r for r in self.pointer_sheet.rows if r[0].value == gamefile.filename]:
            text_offset = int(row[1].value, 16)
            pointer_offset = int(row[2].value, 16)

            # Don't ingest the pointers that point to error messages!
            # They stick around and point to random locations in the spare block,
            # and they mess up the typesetting process.
            if spare_start <= text_offset <= spare_stop:
                continue

            ptr = Pointer(gamefile, pointer_offset, text_offset)

            try:
                # Look for this pointer in the list of pointers to be absorbed.
                #assert pointer_offset != 0x1b6c
                receiver_offset = POINTER_ABSORB[(gamefile.filename, pointer_offset)]

                receiver = ptrs[receiver_offset][0]

                # Edit this pointer to point to the same location as the receiver.
                diff = receiver.text_location - text_offset
                ptr.edit(diff)

                # Now discard the old pointer and create a new one with the updated location.
                ptr = Pointer(gamefile, pointer_offset, receiver.text_location)
                text_offset = receiver.text_location

            except KeyError:
                pass

            try:
                # Look for this pointer in the list of pointers to just... change.
                destination_offset = POINTER_CHANGE[(gamefile.filename, pointer_offset)]
                print "POINTER CHANGED"

                diff = destination_offset - text_offset
                ptr.edit(diff)

                ptr = Pointer(gamefile, pointer_offset, destination_offset)
                text_offset = destination_offset
            except KeyError:
                pass                

            if refresh:
                # Grab the text offset from the filestring instead.
                pointer_word = gamefile.filestring[pointer_offset*2:(pointer_offset+2)*2]
                pointer_value = unpack(pointer_word[:2], pointer_word[2:])
                text_offset = pointer_value + POINTER_CONSTANT[gamefile.filename]

            if text_offset in ptrs:
                ptrs[text_offset].append(ptr)
            else:
                ptrs[text_offset] = [ptr]
        return ptrs


class Overflow(object):
    """A string of data that must be repositioned elsewhere."""
    def __init__(self, gamefile, original_location, bytestring):
        self.gamefile = gamefile
        self.start, self.stop = original_location
        self.original_bytestring = bytestring
        self.bytestring = bytestring

        self.new_length = self.get_length()

    def get_length(self):
        # This duplicates some functionality of move(), because we really just need this to get the new length.
        for offset in range(self.start, self.stop):
            for block in self.gamefile.blocks:
                for trans in [x for x in block.translations if x.location == offset]:

                    jp_bytestring = trans.jp_bytestring
                    en_bytestring = trans.en_bytestring
                    if en_bytestring == "" and trans.english != "[BLANK]":
                        en_bytestring = jp_bytestring
                    try:
                        j = self.bytestring.index(jp_bytestring)
                    except ValueError:
                        print "having problems in the overflow get_length search now"
                        print hex(self.start), hex(self.stop), trans.english
                        #print jp_bytestring
                        jp_bytestring = trans.jp_bytestring_alt
                        #print jp_bytestring
                        #print self.bytestring
                        j = self.bytestring.index(jp_bytestring)
                    self.bytestring = self.bytestring.replace(jp_bytestring, en_bytestring, 1)

        # return the length of the new bytestring, but restore the original bytestring first
        result = len(self.bytestring) // 2
        self._temp_bytestring = self.bytestring
        self.bytestring = str(self.original_bytestring)
        return result

    def move(self, location):
        destination_block = self.gamefile.block_at(location)
        #print "\nmoving", self, "to location", hex(location), "\n"

        pointer_diff = location - self.start

        for pointer in [p for p in self.gamefile.pointers if self.start <= p < self.stop]:
            for i, p in enumerate(self.gamefile.pointers[pointer]):
                # Don't double-edit the text in pointers
                if i < 1:
                    for trans in p.translations:
                        # need to move pointers even when the text remains the same!!
                        # do a loop over pointers instead of translations

                        jp_bytestring = trans.jp_bytestring
                        en_bytestring = trans.en_bytestring

                        if en_bytestring == '' and trans.english != '[BLANK]':
                            en_bytestring = jp_bytestring

                        this_string_diff = (len(en_bytestring) - len(jp_bytestring)) // 2

                        j = self.bytestring.index(jp_bytestring)
                        self.bytestring = self.bytestring.replace(jp_bytestring, en_bytestring, 1)
                        #pointer_diff += this_string_diff

                # When there are multiple translated strings in an overflow bytestring,
                p.edit(pointer_diff)
                #print hex(p.new_text_location)


        assert len(self.bytestring) == len(self._temp_bytestring), "%s:\n%s\n%s" % (self, self.bytestring, self._temp_bytestring)
        destination_block.blockstring += self.bytestring

    def __repr__(self):
        return hex(self.start) + " " + hex(self.stop)
