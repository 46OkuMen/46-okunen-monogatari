"""Hopelessly late 'unit tests' for the reinserter."""
# At this point I need to test various things about the data, not really the program itself.
# So this doesn't really count much for 'test coverage.' Oh well.

from utils import DUMP_XLS, onscreen_length
from openpyxl import load_workbook
from rominfo import CREATURE_BLOCK, CREATURE_MAX_LENGTH, DIALOGUE_MAX_LENGTH, DAT_MAX_LENGTH, FULLSCREEN_MAX_LENGTH

class TestDump:

    def test_increasing_offsets(self):
        """Make sure the offsets are strictly increasing - so that no strings are mislabeled."""
        wb = load_workbook(DUMP_XLS)
        sheets = wb.get_sheet_names()
        sheets.remove('ORIGINAL')
        sheets.remove('MISC TITLES')
        sheets.remove('SINKA.DAT')
        sheets.remove('SEND.DAT')
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            new_offset = 0
            for index, row in enumerate(ws.rows[1:]):  # Skip the first row, it's just labels
                assert int(row[0].value, 16) > new_offset, "In sheet %s, fix offset at row %s; %s < %s" % (sheet, index+2, hex(int(row[0].value, 16)), hex(new_offset))
                new_offset = int(row[0].value, 16)

    def test_all_string_lengths(self):
        """Corasest string test. No string can be longer than 76."""
        wb = load_workbook(DUMP_XLS)
        sheets = wb.get_sheet_names()
        sheets.remove('ORIGINAL')
        sheets.remove('MISC TITLES')
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            new_offset = 0
            for index, row in enumerate(ws.rows[1:]):
                if row[4].value:
                    assert onscreen_length(row[4].value) <= FULLSCREEN_MAX_LENGTH, "%s: %s has length %s" % (sheet, row[4].value, onscreen_length(row[4].value))

    def test_creature_string_lengths(self):
        """No creature name can have a name longer than 21."""
        wb = load_workbook(DUMP_XLS)
        for sheet in ['ST1.EXE', 'ST2.EXE', 'ST3.EXE', 'ST4.EXE', 'ST5.EXE', 'ST6.EXE']:
            creature_lo, creature_hi = CREATURE_BLOCK[sheet]
            ws = wb.get_sheet_by_name(sheet)
            new_offset = 0
            for index, row in enumerate(ws.rows[1:]):
                if int(row[0].value, 16) >= creature_lo and int(row[0].value, 16) <= creature_hi:
                    if row[4].value:
                        assert onscreen_length(row[4].value) <= CREATURE_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)

    def test_dat_string_lengths(self):
        """No encyclopedia string should be longer than 68."""
        wb = load_workbook(DUMP_XLS)
        for sheet in ['SEND.DAT', 'SINKA.DAT']:
            ws = wb.get_sheet_by_name(sheet)
            new_offset = 0
            for index, row in enumerate(ws.rows[1:]):
                if row[4].value:
                    assert onscreen_length(row[4].value) <= DAT_MAX_LENGTH, "In sheet %s, shorten string at row %s" % (sheet, index+2)

# make sure a blank translation sheet doesn't return overflow errors
